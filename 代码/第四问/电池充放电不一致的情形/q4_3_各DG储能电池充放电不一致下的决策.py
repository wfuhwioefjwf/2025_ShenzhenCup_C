#-*- coding: utf-8 -*- 

import math
from numba import njit
from collections import deque
import pandas as pd
import numpy as np
import networkx as nx
from itertools import combinations
import multiprocessing as mp
import os, copy
from typing import Tuple, List, Optional, Dict
import pickle
from pathlib import Path
import itertools
from pathlib import Path
HERE = Path(__file__).resolve().parent

# ============ 常量 ============
EXCEL = str((HERE.parent.parent / "C题附件：增加导线中间的开关和是否为干路两列.xlsx").resolve())
V_BASE = 10e3        
I_RATE = RATED_CURRENT = 220    # 额定电流 单位A
DG_BASE = 300         # DG初始容量 单位kW
TYPE_W = {'居民':0.067,'商业':0.226,'政府和机构':0.354,'办公和建筑':0.354}   # 不同用户权重
main_switches = ['S1','S2','S3','S4','S5','S8','S27','S29','S11','S12'] # 干路开关表

# ==================== 初始图节点列表 ====================
# ==================== 出线开关列表 ====================
circuit_breakers = [
    {'名称': f'CB{i}', '故障状态': False} for i in range(1, 4)  # CB1~CB3
]
# ==================== 分段开关列表 ====================
segment_switches = [
    {'名称': f'S{i}',
     '开关状态': "开",  
     '故障状态': False}
    for i in range(1, 30)  # S1~S29
]
# ==================== 联络开关列表 ====================
tie_switches = [
    {'名称': 'S13-1', '开关状态': '关', '故障状态': False},
    {'名称': 'S29-2', '开关状态': '关', '故障状态': False},
    {'名称': 'S62-3', '开关状态': '关', '故障状态': False}
]
# ==================== 用户列表 ========================
# 类型字典
type_dict = {
    1: '居民',
    2: '商业',
    3: '政府和机构',
    4: '办公和建筑'
}
# 用户类型映射表
user_type_mapping = [
    # 用户1-10
    1, 1, 1, 1, 1, 1, 4, 1, 3, 1,
    # 用户11-20
    2, 4, 1, 4, 1, 2, 1, 4, 1, 1,
    # 用户21-30
    3, 1, 1, 1, 1, 1, 2, 1, 3, 1,
    # 用户31-40
    2, 4, 2, 2, 1, 3, 1, 2, 1, 4,
    # 用户41-50
    1, 2, 1, 1, 3, 1, 4, 1, 2, 1,
    # 用户51-60
    1, 1, 2, 1, 1, 2, 1, 3, 1, 1,
    # 用户61-62
    3, 1
]
users=pd.read_excel(EXCEL,sheet_name=0)
active_power=users['有功P/kW']
customers = [
    {'名称': f'U{i}',
     '类型': type_dict[user_type_mapping[i-1]],
     '用电功率(kW)': active_power[i-1],  
     '故障状态': False}
    for i in range(1, 63) # U1~U62
]
# ==================== 线路列表 =======================
df_topo = pd.read_excel(EXCEL, sheet_name=1, header=0,converters={'导线中间的开关': str,'是否是干路':str})   

# 创建导线编号到起点终点用户的字典
line_dict = {f'L{row["编号"]}': (f'U{row["起点"]}', f'U{row["终点"]}') for _, row in df_topo.iterrows()}

main_lines = df_topo[df_topo['是否是干路'] == 'True']['编号'].apply(lambda x: f'L{x}').tolist()   # 提取干路线路编号（根据"是否是干路"列的值判断）
# 构建完整线路参数
line_parameters = {}
for _, row in df_topo.iterrows():
    line_id   = int(row['编号'])
    length_km = float(row['长度/km'])
    R_raw     = float(row['电阻/Ω'])      
    R_total = R_raw                    
    line_parameters[line_id] = {
        '起点'  : int(row['起点']),
        '终点'  : int(row['终点']),
        '开关'  : str(row['导线中间的开关']),
        '长度(km)' : length_km,
        '总电阻(Ω)' : R_total,                         
        '总电抗(Ω)' : float(row['电抗/Ω']),           
        'main'  : str(row['是否是干路']) == 'True'
    }
lines = [
    {
        '名称'      : f'L{i}',
        '长度(km)'  : line_parameters[i]['长度(km)'],
        '总电阻(Ω)' : line_parameters[i]['总电阻(Ω)'],
        '总电抗(Ω)' : line_parameters[i]['总电抗(Ω)'],
        '总阻抗(Ω)' : complex(
            line_parameters[i]['总电阻(Ω)'],
            line_parameters[i]['总电抗(Ω)']
        ),
        '故障状态'  : False
    }
    for i in range(1, 60) # L1~L59
]

TIE_LINES = [
    ("LT_13_43", 220, 0.01),   # 对应 S13-1
    ("LT_19_29", 220, 0.01),   # 对应 S29-2
    ("LT_23_62", 220, 0.01),   # 对应 S62-3
]
for name, irate, R in TIE_LINES:
    lines.append({
        "名称"      : name,
        "长度(km)"  : 0.01,
        "总电阻(Ω)" : R,
        "总电抗(Ω)" : 0.0,
        "总阻抗(Ω)" : complex(R, 0.0),
        "额定电流(A)": irate,         
        "故障状态"  : False
    })
# ==================== 分布式电源列表 ====================
dgs = [
    {'名称': f'DG{i}',
     '容量(kW)': DG_BASE,  
     '开关状态': '关',
     '故障状态': False}
    for i in range(1, 9)  # DG1~DG8
]

# ==================== 创建网络图 ====================
G = nx.Graph()

# 添加用户节点
for user in customers:
    G.add_node(user['名称'], type='用户', **user)
# 添加导线节点
for line in lines:
    G.add_node(line['名称'], type='导线', **line)
# 添加分段开关
for segment_switch in segment_switches:
    G.add_node(segment_switch['名称'], type='分段开关', **segment_switch)
# 添加联络开关
for tie_switch in tie_switches:
    G.add_node(tie_switch['名称'], type='联络开关', **tie_switch)
# 添加断路器
for breaker in circuit_breakers:
    G.add_node(breaker['名称'], type='断路器', **breaker)
# 添加分布式能源
for dg in dgs:
    G.add_node(dg['名称'], type='分布式能源', **dg)

# ==================== 构建简单边连接 ====================
connections = [
    # 断路器相关连接
    ("CB1", "U1", False),
    ("CB3", "U23", False),
    ("CB2", "U43", False),
    
    # 分布式电源连接
    ("U22", "DG2", False), ("U16", "DG1", False),
    ("U39", "DG5", False), ("U32", "DG3", False),
    ("U35", "DG4", False), ("U52", "DG7", False),
    ("U55", "DG8", False), ("U48", "DG6", False)
]
connections += [
    # S13-1
    ("U13" , "S13-1" , False),
    ("S13-1", "LT_13_43", False),
    ("LT_13_43", "U43" , False),

    # S29-2
    ("U19" , "S29-2" , False),
    ("S29-2", "LT_19_29", False),
    ("LT_19_29", "U29" , False),

    # S62-3
    ("U23" , "S62-3" , False),
    ("S62-3", "LT_23_62", False),
    ("LT_23_62", "U62" , False),
]

# 添加预定义的连接关系
for conn in connections:
    u, v, is_tie = conn
    G.add_edge(u, v, 联络边=is_tie)

# ==================== 处理带分段开关的线路连接 ====================
for line_id, params in line_parameters.items():
    start_user = f"U{params['起点']}"
    end_user = f"U{params['终点']}"
    switch_code = params['开关']  
    line_node = f"L{line_id}"

    # 处理带分段开关的线路
    if switch_code != '0':  
        switch_node = switch_code
        # 构建连接路径：用户-开关-导线-用户
        G.add_edge(start_user, switch_node, 联络边=False)
        G.add_edge(switch_node, line_node, 联络边=False)
        G.add_edge(line_node, end_user, 联络边=False)
    
    # 无分段开关的常规连接
    else: 
        # 构建连接路径：用户-导线-用户
        G.add_edge(start_user, line_node, 联络边=False)
        G.add_edge(line_node, end_user, 联络边=False)

# ==================== 潮流计算 ====================
# 计算干路电流
def feeder_currents(I_amp, net):
    """
    计算三个CB的总电流
    CB1 = L1
    CB2 = L41 + L53 + (L12 if S13-1开启)
    CB3 = L22 + (L59 if S62-3开启)
    """
    # CB1的电流 = L1
    I1 = I_amp.get('L1', 0.0)
    
    # CB2的电流 = L41 + L53
    I2 = I_amp.get('L41', 0.0) + I_amp.get('L53', 0.0)
    # 如果S13-1开启，加上L12
    if net.nodes.get('S13-1', {}).get('开关状态') == '开':
        I2 += I_amp.get('L12', 0.0)
    
    # CB3的电流 = L22
    I3 = I_amp.get('L22', 0.0)
    # 如果S62-3开启，加上L59
    if net.nodes.get('S62-3', {}).get('开关状态') == '开':
        I3 += I_amp.get('L59', 0.0)
    
    return {'CL1': I1, 'CL2': I2, 'CL3': I3}

# 从无向图 G 生成以 CB 为根的径向 DiGraph D：
def build_radial_digraph(G):
    """构建径向有向图用于潮流计算"""
    H = G.copy()

    # 0. 删除故障节点
    fault_nodes = [n for n, d in H.nodes(data=True)
                   if d.get('故障状态', False)
                   and d.get('type') in ('导线', '分段开关', '联络开关', '断路器', '分布式能源')]
    H.remove_nodes_from(fault_nodes)
    
    # 1. 删除"关"的开关
    closed_sw = [n for n, d in H.nodes(data=True)
                 if d.get('type') in ('分段开关', '联络开关')
                 and d.get('开关状态') == '关']
    H.remove_nodes_from(closed_sw)

    # 2. 收缩"开"的开关
    for sw, d in list(H.nodes(data=True)):
        if d.get('type') in ('分段开关', '联络开关') and d.get('开关状态') == '开':
            nbrs = list(H.neighbors(sw))
            if len(nbrs) == 2:
                u, v = nbrs
                H.add_edge(u, v)
                H.add_edge(v, u)
            H.remove_node(sw)

    # 3. 找断路器作为根
    roots = [cb for cb, d in H.nodes(data=True) if d.get('type') == '断路器']
    if not roots:
        raise ValueError("网络中没有断路器节点")

    # 4. 多源BFS
    dist = {}
    queue = deque()
    for cb in roots:
        dist[cb] = 0
        queue.append(cb)

    while queue:
        u = queue.popleft()
        for nbr in H.neighbors(u):
            if nbr not in dist:
                dist[nbr] = dist[u] + 1
                queue.append(nbr)
                
    # 5. 构造有向图
    D = nx.DiGraph()
    for line_id, d in H.nodes(data=True):
        if d.get('type') != '导线':
            continue
        
        ends = [nbr for nbr in H.neighbors(line_id)
                if H.nodes[nbr]['type'] in ('用户', '分布式能源')]
        if len(ends) != 2:
            continue
        
        u, v = ends
        du, dv = dist.get(u, math.inf), dist.get(v, math.inf)
        
        if du < dv:
            head, tail = u, v
        elif dv < du:
            head, tail = v, u
        else:
            head, tail = (u, v) if str(u) < str(v) else (v, u)
        
        # 添加边，包含电阻和电抗
        D.add_edge(head, tail,
                   id=line_id,
                   R=d.get('总电阻(Ω)', 0.0),
                   X=d.get('总电抗(Ω)', 0.0))
    
    return D, roots

def make_nodal_power_dict(G):
    P = {}
    for n, d in G.nodes(data=True):
        if d.get('故障状态', False):
            continue                 # 故障节点不计功率
        if d.get('type') == '用户':
            P[n] = d.get('用电功率(kW)', 0.0)
        elif d.get('type') == '分布式能源' and d.get('开关状态') == '开':
            P[n] = -d.get('容量(kW)', 0.0)
    return P

SQRT3 = math.sqrt(3)
@njit
def _compute_currents(R_arr, P_arr, U_line):
    n = R_arr.shape[0]
    I = np.zeros(n)
    Psend = np.zeros(n)
    for i in range(n):
        Rk = R_arr[i]
        Pk = P_arr[i]
        if Rk > 0.0 and Pk != 0.0:
            disc = U_line*U_line + 4*Rk*Pk*1e3
            I_val = (SQRT3*(math.sqrt(disc)-U_line)) / (6.0*Rk)
            I[i] = I_val if I_val>0.0 else 0.0
        # 计算送端功率
        Psend[i] = SQRT3*U_line*I[i]/1e3
    return Psend, I

def distflow_forward_backward(D, roots, P_nodal, debug=False):
    """
    改进的配电网前推回代潮流计算
    添加收敛判据和迭代控制
    """
    # 获取拓扑排序
    try:
        order = list(nx.topological_sort(D))
    except:
        return {}, {}, {}
    
    order_reverse = list(reversed(order))
    
    # 迭代控制参数
    max_iterations = 50
    tolerance = 1e-6
    
    # 初始化
    V_mag = {n: 1.0 for n in D.nodes()}
    V_mag_prev = {n: 1.0 for n in D.nodes()}  # 保存上一次迭代的电压
    P_flow = {}  # 线路有功功率流(kW)
    Q_flow = {}  # 线路无功功率流(kvar)
    
    # 迭代求解
    for iteration in range(max_iterations):
        # 保存上一次迭代的电压值
        V_mag_prev = V_mag.copy()
        
        # 后推：计算功率流
        P_flow.clear()
        Q_flow.clear()
        
        for node in order_reverse:
            # 节点注入功率
            P_node = P_nodal.get(node, 0.0)  # kW
            Q_node = P_node * math.tan(math.acos(POWER_FACTOR)) if P_node != 0 else 0  # kvar
            
            # 累加流出功率
            P_out = P_node
            Q_out = Q_node
            
            # 加上所有下游线路功率
            for child in D.successors(node):
                edge_data = D.edges[node, child]
                line_id = edge_data['id']
                if line_id in P_flow:
                    P_out += P_flow[line_id]
                    Q_out += Q_flow[line_id]
            
            # 计算流向上游的功率
            for parent in D.predecessors(node):
                edge_data = D.edges[parent, node]
                line_id = edge_data['id']
                R = edge_data['R']
                X = edge_data['X']
                
                # 计算线路损耗
                V_node = V_mag[node] * V_BASE / 1000  # kV
                if V_node > 0:
                    I_approx = math.sqrt(P_out**2 + Q_out**2) / (SQRT3 * V_node)  # A
                    P_loss = 3 * I_approx**2 * R / 1000  # kW
                    Q_loss = 3 * I_approx**2 * X / 1000  # kvar
                else:
                    P_loss = Q_loss = 0
                
                # 送端功率 = 受端功率 + 损耗
                P_flow[line_id] = P_out + P_loss
                Q_flow[line_id] = Q_out + Q_loss
        
        # 前推：更新电压
        for node in order:
            if node in roots:
                V_mag[node] = 1.0
            else:
                for parent in D.predecessors(node):
                    edge_data = D.edges[parent, node]
                    line_id = edge_data['id']
                    R = edge_data['R']
                    X = edge_data['X']
                    
                    P = P_flow.get(line_id, 0)
                    Q = Q_flow.get(line_id, 0)
                    V_parent = V_mag[parent]
                    
                    # 电压降落
                    if V_BASE > 0:
                        delta_V_pu = (P * R + Q * X) / (V_BASE**2 / 1000)
                        V_mag[node] = max(0.8, V_parent - delta_V_pu)  # 限制最低电压
        
        # 检查收敛性
        max_voltage_change = 0.0
        for node in D.nodes():
            voltage_change = abs(V_mag[node] - V_mag_prev[node])
            max_voltage_change = max(max_voltage_change, voltage_change)
        
        if debug and iteration == 0:
            print(f"潮流计算迭代 {iteration + 1}: 最大电压变化 = {max_voltage_change:.6f}")
        
        # 判断是否收敛
        if max_voltage_change < tolerance:
            if debug:
                print(f"潮流计算在第 {iteration + 1} 次迭代后收敛")
            break
    else:
        if debug:
            print(f"潮流计算达到最大迭代次数 {max_iterations}，最大电压变化 = {max_voltage_change:.6f}")
    
    # 计算线路电流
    I_amp = {}
    P_send = {}
    
    for (u, v), data in D.edges.items():
        line_id = data['id']
        if line_id in P_flow:
            P = P_flow[line_id]  # kW
            Q = Q_flow.get(line_id, P * math.tan(math.acos(POWER_FACTOR)))  # kvar
            V_u = V_mag[u] * V_BASE  # V
            
            if V_u > 0:
                # 视在功率
                S = math.sqrt(P**2 + Q**2) * 1000  # VA
                # 三相电流
                I = S / (SQRT3 * V_u)  # A
                
                I_amp[line_id] = I
                P_send[line_id] = P
    
    return P_send, I_amp, V_mag

# 潮流计算主函数入口
POWER_FACTOR = 0.95  # 功率因数
def run_distflow(G):
    """运行潮流计算并检查过载"""
    global fault_counter
    fault_counter = getattr(run_distflow, 'counter', 0) + 1
    run_distflow.counter = fault_counter
    debug = (fault_counter % 100 == 1)  # 每100个故障打印一次调试信息
    
    try:
        D, roots = build_radial_digraph(G)
        if D.number_of_edges() == 0:
            if debug:
                print("无可用导线，无法计算潮流")
            return []
        
        # 构建节点功率字典
        P_nodal = make_nodal_power_dict(G)
        
        # 运行潮流计算
        P_send, I_amp, V_mag = distflow_forward_backward(D, roots, P_nodal)
        
        # 计算CB电流
        cb_currents = feeder_currents(I_amp, G)
        
        # 过载检查
        overload = []
        
        # 线路过载检查
        for line_id, I in I_amp.items():
            if I > 1.1 * RATED_CURRENT:
                overload.append((line_id, I))
        
        # CB过载检查
        for cb_id, I in cb_currents.items():
            I_amp[cb_id] = I  # 保存CB电流到结果中
            if I > 1.1 * RATED_CURRENT:
                overload.append((cb_id, I))
        
        return overload
        
    except Exception as e:
        if debug:
            print(f"潮流计算错误: {e}")
            import traceback
            traceback.print_exc()
        return []

def is_reachable_from_cb(G, cb, target_node):
    powered_areas = []
    if G.nodes[cb].get('故障状态'):
        return False
    visited = []
    stack = [cb]
    while stack:
        node = stack.pop()
        if node in visited:
            continue
        visited.append(node)
        for neighbor in G.neighbors(node):
            if neighbor in visited:
                continue
        # 检查边经过的元件是否故障和开关状态
            if not G.nodes[neighbor].get('故障状态', False):
                if G.nodes[neighbor].get('type') in ['分段开关', '联络开关']:
                    if G.nodes[neighbor].get('开关状态') == '开':
                        stack.append(neighbor)
                else:
                    stack.append(neighbor)
    powered_areas = [n for n in visited if n.startswith('U')]
    if target_node in powered_areas:
        return True
    else:
        return False
    
# 创建NetworkOperator类封装故障相关函数
class NetworkOperator:
    def __init__(self, G):
        self.G = G.copy()
        self.original_G = G
        self._init_network_status()

    # 初始化网络状态（重置并断开联络开关）
    def _init_network_status(self):
        self.G = copy.deepcopy(self.original_G)  
        for n in self.G.nodes:
            if self.G.nodes[n]['type'] == '联络开关':
                self.G.nodes[n]['开关状态'] = '关'
        self._configure_initial_state()

    # 从每个DG出发，搜索到最近的CB并保存路径到visited列表
    def _configure_initial_state(self):
        for dg in dgs:
            visited = []  # 保存所有路径中的节点
            dg_name = dg['名称']
            critical_node=None
            load=0
            total_power=self.G.nodes[dg_name]['容量(kW)']
            if dg_name not in self.G.nodes:
                continue

            # BFS搜索路径
            queue = deque([(dg_name, [dg_name])])
            visited_nodes = set()
            found = False
            while queue and not found:
                current, path = queue.popleft()
                if current in visited_nodes:
                    continue
                visited_nodes.add(current)

                # 检查当前节点是否是CB
                if current.startswith('CB'):
                    visited.extend(path)  # 保存路径节点
                    found = True
                    critical_node = current
                    break

                # 如果当前节点是开关，且开关关闭则阻断路径
                node_data = self.G.nodes[current]
                if node_data['type'] in ['分段开关', '联络开关']:
                    if node_data.get('开关状态', '关') != '开':
                        continue  

                # 遍历邻居节点
                for neighbor in self.G.neighbors(current):
                    if neighbor not in visited_nodes:
                        queue.append((neighbor, path + [neighbor]))

            for node in visited:
                if node.startswith('U'):
                    load += self.G.nodes[node]['用电功率(kW)']
                if load >= total_power and critical_node.startswith('CB'):
                    critical_node = node
                    break
            # 找到关键节点，向 DG 搜索，配置开关和 DG 状态
            if critical_node:
                critical_index = visited.index(critical_node)
                for idx in range(critical_index - 1, -1, -1):
                    node_name = visited[idx]
                    if node_name.startswith('S') and node_name not in main_switches:
                        self.G.nodes[node_name]['开关状态'] = '关'
                        self.G.nodes[dg_name]['开关状态'] = '开'
                        break
    
    # 计算失负荷
    def _risk(self, net, fault_loc=None):

        # ---------- 构造真实连通图 ----------
        H = net.copy()
        remove = []
        for n, d in H.nodes(data=True):
            if d.get('故障状态', False) and d['type'] in {'导线','分段开关','联络开关','断路器'}:
                remove.append(n)
            elif d['type'] in {'分段开关','联络开关'} and d.get('开关状态') == '关':
                remove.append(n)
        H.remove_nodes_from(remove)

        # ---------- 对每个连通分区判断是否"全供" ----------
        lost = 0.0
        for comp in nx.connected_components(H):
            load_sum = 0.0
            gen_cap  = 0.0
            has_cb   = False

            # 统计本分区负荷与发电
            for n in comp:
                nd = H.nodes[n]
                if nd['type'] == '用户':
                    load_sum += nd['用电功率(kW)']
                elif nd['type'] == '断路器' and not nd.get('故障状态', False):
                    has_cb = True           # 视为容量无限
                elif nd['type'] == '分布式能源' and \
                    nd.get('开关状态') == '开' and \
                    not nd.get('故障状态', False):
                    gen_cap += nd['容量(kW)']

            # 是否失电
            if not has_cb:                       
                if gen_cap < load_sum:           
                    # 按每个用户的权重乘以对应的用电功率累加
                    for u in comp:
                        if u.startswith('U'):
                            user_power = H.nodes[u]['用电功率(kW)']
                            user_weight = TYPE_W[H.nodes[u]['类型']]
                            lost += user_power * user_weight

        # ---------- 短路用户单独算 ----------
        if fault_loc and fault_loc.startswith('U'):
            nd = net.nodes[fault_loc]
            lost += nd['用电功率(kW)'] * TYPE_W[nd['类型']]
        return lost

    # 核心算法：根据故障位置 loc 判断故障类型 (1-5)
    def _determine_fault_type(self, loc):
        # 步骤1: 判断是否在干路（直接通过线路或所在线路判断）
        if self._is_main_line_fault(loc, main_lines):
            if loc.startswith('U'):
                return 0
            else:
                return 5  # 类型5，干路故障
        
        # 步骤2: 如果不在干路上，向两边搜索直到碰到干路或 DG
        is_found_dg = False
        found_dg = None
        for direction in ['upstream', 'downstream']:
            path = self._search_until_main_line_or_dg(loc, direction)
            
            # 如果路径中找到 DG，则标记并结束搜索
            for node in path:
                if self.G.nodes[node]['type'] == '分布式能源':
                    is_found_dg = True
                    found_dg = node
                    break
            
            if is_found_dg:
                break
        
        # 如果没找到 DG，则按情况2处理
        if not is_found_dg:
            return 2  # 类型2，未找到DG，视为支路故障

        # 步骤3: DG 是打开的（DG 自己的开关状态 == '开'）
        if self.G.nodes[found_dg]['开关状态'] == '开':
            if self._has_switch_in_path(loc, found_dg, '关'):
                return 2      # 类型 2：同一方向上有断开的开关
            else:
                return 1      # 类型 1：这条方向上一路畅通

        # 步骤4: DG 是关闭的（DG 自己的开关状态 == '关'）
        if self.G.nodes[found_dg]['开关状态'] == '关':
            if self._has_switch_in_path(loc, found_dg, '开'):
                return 4      # 类型 4：这条方向上有开启的开关
            else:
                return 3      # 类型 3：这条方向上完全没有开关
            
        return 5  # 其余默认归为干路故障

    def _has_switch_in_path(self, start, end, target_state):
        try:
            path = nx.shortest_path(self.G, start, end)
        except nx.NetworkXNoPath:
            return False

        for node in path:
            if self.G.nodes[node]['type'] in ('分段开关', '联络开关'):
                if self.G.nodes[node]['开关状态'] == target_state:
                    return True
        return False
        
    def _is_main_line_fault(self, loc, main_lines):
        # 情况1: 故障位置是干路线路
        if loc.startswith('L'):
            return loc in main_lines

        # 情况2: 故障位置是用户，检查其连接线路是否为干路
        if loc.startswith('U'):
            connected_lines = []
            for nbr in self.G.neighbors(loc):
                if self.G.nodes[nbr]['type'] == '导线':
                    connected_lines.append(nbr)

            # 检查是否所有连接的线路中是否有干路
            if len(connected_lines) > 0:
                return any(line in main_lines for line in connected_lines)
            return False
        if loc.startswith('CB'):
            return True
        if loc in ['S13-1', 'S29-2', 'S62-3']:
            return True
        else:
            return loc in main_switches
    
    def _search_until_main_line_or_dg(self, loc, direction):
        path = []
        current_node = loc
        visited = set()  # 记录已访问节点
        
        while True:
            if current_node in visited:
                break
            visited.add(current_node)
            
            neighbors = list(self.G.neighbors(current_node))
            if direction == 'upstream':
                neighbors = list(reversed(neighbors))
                
            found_target = False
            for nbr in neighbors:
                if nbr in visited:
                    continue
                    
                if nbr.startswith('L') and nbr in main_lines:
                    path.append(nbr)
                    found_target = True
                    break
                if self.G.nodes[nbr].get('type') == '分布式能源':
                    path.append(nbr)
                    found_target = True
                    break
            
            if found_target:
                break
            else:
                valid_neighbors = [n for n in neighbors if n not in visited]
                if not valid_neighbors:
                    break
                    
                path.append(current_node)
                current_node = valid_neighbors[0]  # 取第一个有效邻居
        
        return path

    def _has_open_switch_in_path(self, start, end):
        try:
            path = nx.shortest_path(self.G, start, end)
        except nx.NetworkXNoPath:
            return False

        for node in path:
            if self.G.nodes[node]['type'] in ['分段开关', '联络开关'] and self.G.nodes[node]['开关状态'] == '关':
                return True
        return False

    def _judge_main_fault_type(self, loc):
        tie_switch_constraints = {}
        if loc in ['CB1', 'L1', 'L2', 'L3', 'S1']:
            tie_switch_constraints = {
                "S62-3": ["S29", "S27"]                       # 打开S62-3时，S29/S27至少关一个
            }
        elif loc in ['L4', 'L5', 'L6', 'L7', 'L8', 'L9', 'L10', 'L11', 'L12',
                        'S2', 'S3', 'S4', 'S5']:
            tie_switch_constraints = {
                "S29-2": ["S1", "S8", "S11", "S12"],          # 打开S29-2时，S1/S8/S11/S12至少关一个
                "S62-3": ["S29", "S27"]                       # 打开S62-3时，S29/S27至少关一个
            }
        elif loc in ['CB2']:
            tie_switch_constraints = {
                "S29-2": ["S1", "S8", "S11", "S12"],          # 打开S29-2时，S1/S8/S11/S12至少关一个
            }
        elif loc in ['L53', 'L54', 'L57', 'L58', 'L59', 'S27', 'S29']:
            tie_switch_constraints = {
                "S13-1": ["S1", "S2", "S3", "S4", "S5"],      # 打开S13-1时，S1/S2/S3/S4/S5至少关一个
                "S29-2": ["S1", "S8", "S11", "S12"],          # 打开S29-2时，S1/S8/S11/S12至少关一个
            }
        elif loc in ['CB3']:
            tie_switch_constraints = {
                "S13-1": ["S1", "S2", "S3", "S4", "S5"],      # 打开S13-1时，S1/S2/S3/S4/S5至少关一个
            }
        elif loc in ['L22', 'L23', 'L24', 'L25', 'L26', 'L27', 'S11', 'S12',
                        'L16', 'L17', 'L18', 'S8']:
            tie_switch_constraints = {
                "S13-1": ["S1", "S2", "S3", "S4", "S5"],      # 打开S13-1时，S1/S2/S3/S4/S5至少关一个
                "S62-3": ["S29", "S27"]                       # 打开S62-3时，S29/S27至少关一个
            }
        return tie_switch_constraints

    # 单故障核心处理函数，返回损失、策略等
    def simulate_fault(self, ftype, loc, reset=True):
        if reset:
            self._init_network_status()

        # 1 获取网络状态
        branch_dg_list = self._branch_dgs(loc, self.G)

        # 2 计算用户短路额外损失
        node_data = self.G.nodes[loc]
        if node_data.get('type') == '用户':
            extra_lost = node_data['用电功率(kW)'] * TYPE_W[node_data['类型']]
        else:
            extra_lost = 0

        # 3 构建基础故障网络
        net0 = copy.deepcopy(self.G)
        net0.nodes[loc]['故障状态'] = True
        if loc.startswith('L'):
            base_label = f"断路线路 {loc}"
        elif loc.startswith('S') or loc.startswith('CB'):     
            net0.nodes[loc]['开关状态'] = '关'     
            base_label = f"跳闸 {loc}"
        elif loc.startswith('DG'):    
            net0.nodes[loc]['开关状态'] = '关'     
            base_label = f"{loc}故障"
        else:
            lines = [n for n in net0.neighbors(loc)]
            if len(lines) >= 2:
                u, v = lines[0], lines[1]
                net0.add_edge(u, v, 联络边=False)
            base_label = f"短路用户 {loc}"

        # 4 候选策略生成
        candidates = []
        # 公共基础策略
        candidates.append((net0, base_label)) 

        # 故障类型1：打开上游开关
        if ftype == 1:
            # 向上游 BFS 找最近的分段/联络开关
            visited, queue = set(), deque([loc])
            found_switch = None
            while queue:
                curr = queue.popleft()
                if curr in visited:
                    continue
                visited.add(curr)

                if net0.nodes[curr]['type'] in ('分段开关', '联络开关') and net0.nodes[curr]['开关状态'] == '关':
                    found_switch = curr
                    break
                queue.extend([nbr for nbr in net0.neighbors(curr) if nbr not in visited])
            new_net = copy.deepcopy(net0)
            new_net.nodes[found_switch]['开关状态'] = '开'

            # 考虑接入原本由DG供应的用户后可能会出现过负荷的情形
            overload = run_distflow(new_net)
            if not overload:
                candidates.append((new_net, f"打开开关 {found_switch}"))
            else:
                # 添加原始策略
                candidates.append((new_net, f"打开开关 {found_switch}"))
                # 过负荷的情况，添加联络线策略
                switches = ['S8', 'S11', 'S12', 'S1', 'S2', 'S3', 'S4', 'S5', 'S29', 'S27', 'S62-3', 'S13-1', 'S29-2']
                max_switch_num = 4  # 最大同时操作开关数量
                key_ties = {'S13-1', 'S29-2', 'S62-3'}

                # 定义联络开关及其约束关系
                tie_switch_constraints = {}
                tie_switch_constraints = self._judge_main_fault_type(loc)

                for i in range(1, min(max_switch_num, len(switches)) + 1):
                    for combo in combinations(switches, i):
                        # 判断是否包含至少一个关键联络开关
                        if not key_ties.intersection(combo):
                            continue
                        new_net = copy.deepcopy(net0)
                        label_parts = [f"打开开关{found_switch}"]
                        for sw in combo:
                            if sw in new_net.nodes:  
                                if new_net.nodes[sw]['开关状态'] == '关':
                                    new_net.nodes[sw]['开关状态'] = '开'
                                    label_parts.append(f"开{sw}")
                                else:
                                    new_net.nodes[sw]['开关状态'] = '关'
                                    label_parts.append(f"关{sw}")
                        valid = True
                        for sw in tie_switch_constraints:
                            if sw in new_net.nodes and new_net.nodes[sw]['开关状态'] == '开':
                                if not any(new_net.nodes[n]['开关状态'] == '关' for n in tie_switch_constraints[sw]):
                                    valid = False
                                    break
                        if valid:
                            candidates.append((new_net, "操作：" + "+".join(label_parts)))
                
        # 故障类型5：组合开关操作
        elif ftype == 5:
            switches = ['S8', 'S11', 'S12', 'S1', 'S2', 'S3', 'S4', 'S5', 'S29', 'S27', 'S62-3', 'S13-1', 'S29-2']
            max_switch_num = 4 # 最大同时操作开关数量

            # 定义联络开关及其约束关系
            tie_switch_constraints = {}
            tie_switch_constraints = self._judge_main_fault_type(loc)

            for i in range(1, min(max_switch_num, len(switches)) + 1):
                for combo in combinations(switches, i):
                    new_net = copy.deepcopy(net0)
                    label_parts = []
                    for sw in combo:
                        if sw in new_net.nodes:  
                            if new_net.nodes[sw]['开关状态'] == '关':
                                new_net.nodes[sw]['开关状态'] = '开'
                                label_parts.append(f"开{sw}")
                            else:
                                new_net.nodes[sw]['开关状态'] = '关'
                                label_parts.append(f"关{sw}")
                    valid = True
                    for sw in tie_switch_constraints:
                        if sw in new_net.nodes and new_net.nodes[sw]['开关状态'] == '开':
                            if not any(new_net.nodes[n]['开关状态'] == '关' for n in tie_switch_constraints[sw]):
                                valid = False
                                break
                    if valid:
                        candidates.append((new_net, "操作：" + "+".join(label_parts)))
            
        # 故障类型3：断开分支DG
        elif ftype == 3:
            for dg in branch_dg_list:
                if dg not in net0.nodes:
                    continue

                # (1) 向上游 BFS 找最近的分段/联络开关
                visited, queue = set(), deque([dg])
                found_switch = None
                while queue:
                    curr = queue.popleft()
                    if curr in visited:
                        continue
                    visited.add(curr)

                    if net0.nodes[curr]['type'] in ('分段开关', '联络开关'):
                        found_switch = curr
                        break
                    queue.extend([nbr for nbr in net0.neighbors(curr) if nbr not in visited])

                if not found_switch:
                    continue 

                # (2) 生成临时网络：开 DG，关最近开关
                tmp_net = copy.deepcopy(net0)
                tmp_net.nodes[dg]['开关状态'] = '开'
                if not loc.startswith('L'):
                    tmp_net.nodes[found_switch]['开关状态'] = '关'

                # (3) 删除该开关节点，计算 DG 可到达区域
                H = tmp_net.copy()
                H.remove_node(found_switch)
                reachable = nx.node_connected_component(H, dg)

                # (4) 统计区域内用户负荷（不含故障节点）
                load_sum = 0.0
                for n in reachable:
                    n_data = tmp_net.nodes[n]
                    if n_data['type'] == '用户' and n != loc:
                        load_sum += n_data['用电功率(kW)']

                # (5) 若 DG 能覆盖该负荷，才加入候选策略
                if load_sum <= tmp_net.nodes[dg]['容量(kW)']:
                    candidates.append((tmp_net, f"开{dg} + 断{found_switch}"))

        # 其他类型仅保留基础策略
        else:
            pass

        # 5 评估所有候选策略
        best_risk      = (float('inf'), float('inf'), float('inf'))
        best_strategy  = base_label
        best_overload  = []         

        for net, label in candidates:
            # 检查各 CB 送电区域是否超额
            for cb in ['CB1', 'CB2', 'CB3']:
                if cb not in net.nodes:
                    continue
                if net.nodes[cb].get('故障状态', False):
                    continue

                # 找 CB 可达区域
                visited = set()
                queue = deque([cb])
                total_load = 0.0

                while queue:
                    node = queue.popleft()
                    if node in visited:
                        continue
                    visited.add(node)
                    for nbr in net.neighbors(node):
                        if nbr in visited:
                            continue
                        if net.nodes[nbr].get('故障状态', False):
                            continue
                        if net.nodes[nbr]['type'] in ['分段开关', '联络开关']:
                            if net.nodes[nbr]['开关状态'] == '开':
                                queue.append(nbr)
                        else:
                            queue.append(nbr)

                # 汇总负荷
                for n in visited:
                    n_data = net.nodes[n]
                    if n_data['type'] == '用户':
                        total_load += n_data.get('用电功率(kW)', 0.0)

            # 计算潮流
            overload = run_distflow(net)
            w1 = 10      # 1.1~1.3 倍区间线性权重
            w2 = 10     # 超过 1.3 倍后二次权重
            threshold = 1.3
            over_penalty = 0.0

            for L, I in overload:
                ratio = I / I_RATE

                # 计算不同的过负荷线路影响的用户权重
                w_0 = 0
                if L in line_dict:
                    U1, U2 = line_dict[L]
                    U1_power = net.nodes[U1]['用电功率(kW)']
                    U2_power = net.nodes[U2]['用电功率(kW)']
                    U1_weight = TYPE_W[net.nodes[U1]['类型']]
                    U2_weight = TYPE_W[net.nodes[U2]['类型']]
                    w_0 = U1_weight * U1_power + U2_weight * U2_power
                elif L.startswith('CL'):
                    if L == 'CL1':
                        U = 'U1'
                    elif L == 'CL2':
                        U = 'U43'
                    elif L == 'CL3':
                        U = 'U23'
                    U_power = net.nodes[U]['用电功率(kW)']
                    U_weight = TYPE_W[net.nodes[U]['类型']]
                    w_0 = U_weight * U_power
                elif L.startswith('LT'):
                    if L == 'LT_13_43':
                        U1 = 'U13'
                        U2 = 'U43'
                    elif L == 'LT_19_29':
                        U1 = 'U19'
                        U2 = 'U29'
                    elif L == 'LT_23_62':
                        U1 = 'U23'
                        U2 = 'U62'
                    U1_power = net.nodes[U1]['用电功率(kW)']
                    U2_power = net.nodes[U2]['用电功率(kW)']
                    U1_weight = TYPE_W[net.nodes[U1]['类型']]
                    U2_weight = TYPE_W[net.nodes[U2]['类型']]
                    w_0 = U1_weight * U1_power + U2_weight * U2_power

                if ratio <= threshold:
                    # 1.1→1.3 线性累加
                    over_penalty += w_0 * w1 * (ratio - 1.1) 
                else:
                    # 先把 >1.1 的线性罚分
                    over_penalty += w_0 * w1 * (ratio - 1.1) 
                    # 再把 >1.3 的部分二次放大
                    over_penalty += w_0 * w2 * (1 +(ratio - threshold) ) ** 2 
            
            net2 = copy.deepcopy(net)
            lost = self._risk(net2, loc)
            total =lost + over_penalty 
            # 结果乘以1e4，保持数据精度
            total = total * 1e4
            lost = lost * 1e4
            over_penalty = over_penalty * 1e4

            # 更新最优策略
            if total < best_risk[2] or (total == best_risk[2] and over_penalty < best_risk[1]):
                best_risk     = (lost, over_penalty, total)
                best_strategy = label
                best_overload = overload         

        return (*best_risk, best_strategy, best_overload)

    def _branch_dgs(self,node,net):
        comp = nx.node_connected_component(net,node)
        return [u for u in comp if net.nodes[u]['type']=='分布式能源']

    # 用于在 (1,1) 场景下判断哪一个故障更靠近主干
    def _distance_to_main(self, loc, net):
        
        cb = self._find_feeder_root(loc, net)
        if cb is None:
            return float('inf')
        try:
            return nx.shortest_path_length(net, source=loc, target=cb)
        except nx.NetworkXNoPath:
            return float('inf')

    # 单点故障 loc 的候选操作生成器
    def _generate_candidates_for_fault(self, loc):
        
        # 先打故障标记在一个临时网络上（但不改开关状态）
        net0 = copy.deepcopy(self.original_G)
        net0.nodes[loc]['故障状态'] = True

        ftype = self._determine_fault_type(loc)
        cands = []

        # 基础“不操作”策略
        cands.append(([], f"不操作（{loc} 基本策略）"))

        # 类型1：打开上游的断开开关
        if ftype == 1:
            # 向上游 BFS 找最近的分段/联络开关
            visited, queue = set(), deque([loc])
            found_switch = None
            while queue:
                curr = queue.popleft()
                if curr in visited:
                    continue
                visited.add(curr)

                if net0.nodes[curr]['type'] in ('分段开关', '联络开关') and net0.nodes[curr]['开关状态'] == '关':
                    found_switch = curr
                    break
                queue.extend([nbr for nbr in net0.neighbors(curr) if nbr not in visited])
            new_net = copy.deepcopy(net0)
            new_net.nodes[found_switch]['开关状态'] = '开'

            # 考虑接入原本由DG供应的用户后可能会出现过负荷的情形
            overload = run_distflow(new_net)
            if not overload:
                cands.append(( [("open", found_switch)], f"打开开关 {found_switch}" ))
            else:
                # 添加原始策略
                cands.append(( [("open", found_switch)], f"打开开关 {found_switch}" ))
                # 过负荷的情况，添加联络线策略
                switches = ['S8', 'S11', 'S12', 'S1', 'S2', 'S3', 'S4', 'S5', 'S29', 'S27', 'S62-3', 'S13-1', 'S29-2']
                max_switch_num = 3  # 最大同时操作开关数量
                key_ties = {'S13-1', 'S29-2', 'S62-3'}
                # 定义联络开关及其约束关系
                tie_switch_constraints = {
                }
                for i in range(1, min(max_switch_num, len(switches)) + 1):
                    for combo in combinations(switches, i):
                        # 判断是否包含至少一个关键联络开关
                        if not key_ties.intersection(combo):
                            continue
                        new_net = copy.deepcopy(net0)
                        label_parts = [f"打开开关{found_switch}"]
                        ops = [("open", found_switch)]
                        for sw in combo:
                            if sw in new_net.nodes:  
                                if new_net.nodes[sw]['开关状态'] == '关':
                                    new_net.nodes[sw]['开关状态'] = '开'
                                    label_parts.append(f"开{sw}")
                                    ops.append(("open", sw))
                                else:
                                    new_net.nodes[sw]['开关状态'] = '关'
                                    label_parts.append(f"关{sw}")
                                    ops.append(("close", sw))
                        # 检查是否符合约束关系
                        valid = True
                        for sw in combo:
                            if sw in tie_switch_constraints and new_net.nodes[sw]['开关状态'] == '开':
                                # 检查约束开关是否至少有一个关
                                if not any(new_net.nodes[const_sw]['开关状态'] == '关' for const_sw in tie_switch_constraints[sw]):
                                    valid = False
                                    break
                        if valid:
                            cands.append((ops, "操作：" + "+".join(label_parts)))

        # 类型5：组合开关操作
        elif ftype == 5:
            switches = ['S8', 'S11', 'S12', 'S1', 'S2', 'S3', 'S4', 'S5', 'S29', 'S27', 'S62-3', 'S13-1', 'S29-2']
            max_switch_num = 3  # 最大同时操作开关数量
            # 定义联络开关及其约束关系
            tie_switch_constraints = {
            }
            for i in range(1, min(max_switch_num, len(switches)) + 1):
                for combo in combinations(switches, i):
                    new_net = copy.deepcopy(net0)
                    label_parts = []
                    ops = []
                    for sw in combo:
                        if sw in new_net.nodes:  
                            if new_net.nodes[sw]['开关状态'] == '关':
                                new_net.nodes[sw]['开关状态'] = '开'
                                label_parts.append(f"开{sw}")
                                ops.append(("open", sw))
                            else:
                                new_net.nodes[sw]['开关状态'] = '关'
                                label_parts.append(f"关{sw}")
                                ops.append(("close", sw))
                    # 检查是否符合约束关系
                    valid = True
                    for sw in combo:
                        if sw in tie_switch_constraints and new_net.nodes[sw]['开关状态'] == '开':
                            # 检查约束开关是否至少有一个关
                            if not any(new_net.nodes[const_sw]['开关状态'] == '关' for const_sw in tie_switch_constraints[sw]):
                                valid = False
                                break
                    if valid:
                        cands.append((ops, "操作：" + "+".join(label_parts)))

        # 类型3：断开分支 DG
        elif ftype == 3:
            branch_dg_list = self._branch_dgs(loc, self.original_G)
            for dg in branch_dg_list:
                # 找到距离 dg 最近的分段/联络开关
                visited, queue = set(), deque([dg])
                found_sw = None
                while queue and found_sw is None:
                    u = queue.popleft()
                    visited.add(u)
                    for nbr in self.original_G.neighbors(u):
                        if nbr in visited: continue
                        if self.original_G.nodes[nbr]['type'] in ('分段开关','联络开关'):
                            found_sw = nbr
                            break
                        queue.append(nbr)
                if not found_sw: 
                    continue
                # 把 DG 打开、该开关关闭
                ops = [("open", dg), ("close", found_sw)]
                label = f"开{dg} + 断{found_sw}"
                cands.append((ops, label))

        # 其余类型仅保留“不操作”
        return cands
    
    # 寻找 loc 由哪个 CB 供电
    def _find_feeder_root(self, loc, net):
        for cb in ['CB1','CB2','CB3']:
            if not net.nodes[cb].get('故障状态', False) \
               and is_reachable_from_cb(net, cb, loc):
                return cb
        return None

    # 判断 loc1, loc2 是否在同一支路上
    def _is_same_branch(self, loc1, loc2, net):
        mains = main_lines
        def attached_main(loc):
            for nbr in net.neighbors(loc):
                if nbr in mains:
                    return nbr
            return None

        m1 = attached_main(loc1)
        m2 = attached_main(loc2)
        return (m1 is not None) and (m1 == m2)

    # 双故障处理核心函数，同样返回损失、策略等
    def simulate_two_faults(self, loc1, loc2):

        # ---------- 构造“基础双故障网络” -----------------
        net0 = copy.deepcopy(self.original_G)
        def _apply_single_fault(g, loc):
            if loc not in g.nodes:
                return
            g.nodes[loc]['故障状态'] = True
            if loc.startswith('L'):                     
                g.remove_node(loc)
            elif loc.startswith(('S', 'CB', 'DG')):           
                g.nodes[loc]['开关状态'] = '关'

        _apply_single_fault(net0, loc1)
        _apply_single_fault(net0, loc2)

        # ---------- 生成候选操作组合 -----------------------
        f1 = self._determine_fault_type(loc1)
        f2 = self._determine_fault_type(loc2)
        combined = []          

        # 1 判断馈线 / 支路关系
        feeder1 = self._find_feeder_root(loc1, self.original_G)
        feeder2 = self._find_feeder_root(loc2, self.original_G)
        same_feeder  = (feeder1 == feeder2)
        same_branch  = same_feeder and self._is_same_branch(loc1, loc2,
                                                            self.original_G)
        # 2 获取各自的单故障候选 
        c1 = self._generate_candidates_for_fault(loc1)
        c2 = self._generate_candidates_for_fault(loc2)

        if f1 == f2 == 5:   # 两个都是干路故障，只针对更靠近主干的做控制
            nearer = min((loc1, loc2),
                         key=lambda x: self._distance_to_main(x,
                                                              self.original_G))
            cs = self._generate_candidates_for_fault(nearer)
            for ops, lbl in cs:
                combined.append((ops, [], lbl+f"(源:{nearer})"))

        elif not same_feeder or (same_feeder and not same_branch):
            # 不同馈线或同馈线不同支路 → 做笛卡儿积
            for ops1, lbl1 in c1:
                for ops2, lbl2 in c2:
                    combined.append((ops1, ops2, lbl1 + " + " + lbl2))
        else:
            # 同支路 → 四种特殊规则
            if f1 == f2 == 1:
                nearer = min((loc1, loc2),
                             key=lambda x: self._distance_to_main(x,
                                                                  self.original_G))
                for ops, lbl in self._generate_candidates_for_fault(nearer):
                    combined.append((ops, [], lbl))
            elif {f1, f2} == {1, 2}:
                combined.append(([], [], "不做操作"))
            elif f1 == f2 == 3:
                for loc in (loc1, loc2):
                    for ops, lbl in self._generate_candidates_for_fault(loc):
                        combined.append((ops, [], lbl))
            elif (f1, f2).count(4) and (f1, f2).count(3):
                tgt = loc1 if f1 == 4 else loc2
                for ops, lbl in self._generate_candidates_for_fault(tgt):
                    combined.append((ops, [], lbl))
            else:
                combined.append(([], [], "不做操作"))
        
        tie_switch_constraints = {}
        tie_switch_constraints_1  = {
            "S13-1": ["S1", "S2", "S3", "S4", "S5"],      # 打开S13-1时，S1/S2/S3/S4/S5至少关一个
            "S29-2": ["S1", "S8", "S11", "S12"],          # 打开S29-2时，S1/S8/S11/S12至少关一个
            "S62-3": ["S29", "S27"]                       # 打开S62-3时，S29/S27至少关一个
        }
        tie_switch_constraints_2 = {
            "S13-1": ["S1", "S2", "S3", "S4", "S5"],      # 打开S13-1时，S1/S2/S3/S4/S5至少关一个
            "S29-2": ["S1", "S8", "S11", "S12"],          # 打开S29-2时，S1/S8/S11/S12至少关一个
            "S62-3": ["S29", "S27"]                       # 打开S62-3时，S29/S27至少关一个
        }
        # 故障类型为 1 或 5 时需要考虑断路
        if f1 == 1 or f1 == 5:
            tie_switch_constraints_1 = self._judge_main_fault_type(loc1)
        if f2 == 1 or f2 == 5:
            tie_switch_constraints_2 = self._judge_main_fault_type(loc2)

        # 求交集
        tie_switch_constraints = {k: set(tie_switch_constraints_1[k]) & set(tie_switch_constraints_2[k]) 
                                    for k in tie_switch_constraints_1.keys() & tie_switch_constraints_2.keys() 
                                    if set(tie_switch_constraints_1[k]) & set(tie_switch_constraints_2[k])}
        
        # 检查是否符合约束关系
        filtered_combined = []
        for combo, ops, label_parts in combined:
            valid = True
            for key, constraint_set in tie_switch_constraints.items():
                # 如果key在combo中，则其约束开关至少有一个也在combo中
                if key in combo:
                    if not any(const_sw in combo for const_sw in constraint_set):
                        valid = False
                        break
            if valid:
                filtered_combined.append((combo, ops, label_parts))
        combined = filtered_combined

        # ---------- 评估所有候选 ----------
        def _merge_ops(ops1, ops2):
            d = {}
            for act, sw in ops1: d[sw] = act
            for act, sw in ops2: d[sw] = act
            return list(d.items())

        best = (float('inf'), float('inf'), float('inf'), "无策略")
        for ops1, ops2, label in combined:
            net = copy.deepcopy(net0)           
            for sw, act in _merge_ops(ops1, ops2):
                if sw in net.nodes:
                    net.nodes[sw]['开关状态'] = '开' if act == 'open' else '关'

            # 计算潮流
            overload = run_distflow(net)
            w1 = 10      # 1.1~1.3 倍区间线性权重
            w2 = 10     # 超过 1.3 倍后二次权重
            threshold = 1.3
            over_penalty = 0.0

            for L, I in overload:
                ratio = I / I_RATE

                # 计算不同的过负荷线路影响的用户权重
                w_0 = 0
                if L in line_dict:
                    U1, U2 = line_dict[L]
                    U1_power = net.nodes[U1]['用电功率(kW)']
                    U2_power = net.nodes[U2]['用电功率(kW)']
                    U1_weight = TYPE_W[net.nodes[U1]['类型']]
                    U2_weight = TYPE_W[net.nodes[U2]['类型']]
                    w_0 = U1_weight * U1_power + U2_weight * U2_power
                elif L.startswith('CL'):
                    if L == 'CL1':
                        U = 'U1'
                    elif L == 'CL2':
                        U = 'U43'
                    elif L == 'CL3':
                        U = 'U23'
                    U_power = net.nodes[U]['用电功率(kW)']
                    U_weight = TYPE_W[net.nodes[U]['类型']]
                    w_0 = U_weight * U_power
                elif L.startswith('LT'):
                    if L == 'LT_13_43':
                        U1 = 'U13'
                        U2 = 'U43'
                    elif L == 'LT_19_29':
                        U1 = 'U19'
                        U2 = 'U29'
                    elif L == 'LT_23_62':
                        U1 = 'U23'
                        U2 = 'U62'
                    U1_power = net.nodes[U1]['用电功率(kW)']
                    U2_power = net.nodes[U2]['用电功率(kW)']
                    U1_weight = TYPE_W[net.nodes[U1]['类型']]
                    U2_weight = TYPE_W[net.nodes[U2]['类型']]
                    w_0 = U1_weight * U1_power + U2_weight * U2_power
                
                if ratio <= threshold:
                    # 1.1→1.3 线性累加
                    over_penalty += w_0 * w1 * (ratio - 1.1) 
                else:
                    # 先把 >1.1 的线性罚分
                    over_penalty += w_0 * w1 * (ratio - 1.1) 
                    # 再把 >1.3 的部分二次放大
                    over_penalty += w_0 * w2 * (1 +(ratio - threshold) ) ** 2

            # 失负荷  
            lost = self._risk(net) + self._extra_user_loss((loc1, loc2))
            total = (lost + over_penalty)

            lost = lost * 1e4
            over_penalty = over_penalty * 1e4
            total = total * 1e4

            if total < best[2]:
                best = (lost, over_penalty, total, label)

        return best

    # 用户短路附加损失
    def _extra_user_loss(self, loc_pair):
        extra = 0.0
        for loc in loc_pair:
            if loc.startswith('U') and loc in self.original_G.nodes:
                nd = self.original_G.nodes[loc]
                extra += nd['用电功率(kW)'] * TYPE_W[nd['类型']]
        return extra

# 对故障节点分类
def device_category(loc):
    if loc.startswith('CB') or loc.startswith('S'):
        return 'switch'
    elif loc.startswith('DG'):
        return 'dg'
    elif loc.startswith('L'):
        return 'line'
    else:
        return 'other'

# 初始化网络
def init_worker(base_net, p_cond, prod_safe):
    global TEMPLATE_G, op, P_COND, PROD_SAFE
    TEMPLATE_G   = base_net
    op           = NetworkOperator(base_net)
    P_COND       = p_cond          
    PROD_SAFE    = prod_safe       

# simulate_single：simulate_fault接口函数
def simulate_single(loc):
    _reset_graph()

    ftype = op._determine_fault_type(loc)
    lost, over, tot, strat, overload = op.simulate_fault(
        ftype, loc, reset=False)
    w = P_COND[loc]      # 计算概率         
    return loc, lost, over, tot, strat, overload, w

# simulate_double：simulate_two_faults接口函数
def simulate_double(pair):
    loc1, loc2 = pair
    _reset_graph()

    lost, over, tot, strat = op.simulate_two_faults(loc1, loc2)
    w = P_COND[loc1] * P_COND[loc2] / PROD_SAFE     # 计算概率  
    return loc1, loc2, lost, over, tot, strat, w

# 重置网络
def _reset_graph():
    g = op.G
    for n, d0 in TEMPLATE_G.nodes(data=True):
        d  = g.nodes[n]
        d.clear();   d.update(d0)          
    g.remove_edges_from(list(g.edges))      
    g.add_edges_from(TEMPLATE_G.edges(data=True))

def process_capacity(DG_capacity):
    """
    只传入DG_capacity（列表或一维数组），其余参数均在函数内推断。
    完成一次基准网络初始化、概率计算、并行计算，并输出Excel结果（若已移除则仅打印累计）。
    """
    # 推断需要的常量
    cap = DG_capacity[0] if isinstance(DG_capacity, (list, np.ndarray)) else DG_capacity
    BASE_XLSX_DIR = r"C:\Users\xueyixian\Desktop\深圳杯\第二问\问题二_8.16_测试"
    PROGRESS_INT = 100
    cpu_num = max(1, mp.cpu_count() - 1)

    # 生成一次已初始化好的基准网络BASE_NET
    G_cap = copy.deepcopy(G)
    dg_nodes = [n for n, d in G.nodes(data=True) if d.get('type') == '分布式能源']
    for dg, dg_cap in zip(dg_nodes, DG_capacity):
        G_cap.nodes[dg]['容量(kW)'] = dg_cap
    op0 = NetworkOperator(G_cap)
    BASE_NET = copy.deepcopy(op0.G)

    # 初始故障概率
    p_fail = {}
    for n, d in op0.G.nodes(data=True):
        typ = d['type']
        if typ == '导线':
            p_fail[n] = d['长度(km)'] * 0.002
        elif typ in ('用户', '分布式能源'):
            p_fail[n] = 0.005
        elif typ in ('分段开关', '联络开关', '断路器'):
            p_fail[n] = 0.002
        else:
            p_fail[n] = 0.0
    p_safe = {n: 1.0 - p for n, p in p_fail.items()}

    all_faults = (
          [f"L{i}"  for i in range(1, 60)]
        + [f"U{i}"  for i in range(1, 63)]
        + [f"S{i}"  for i in range(1, 30)] + ["S13-1", "S29-2", "S62-3"]
        + ["CB1", "CB2", "CB3"]
        + [f"DG{i}" for i in range(1, 9)]
    )
    devices_by_category = {}
    for j in all_faults:
        if j in op0.G.nodes:
            category = device_category(j)
            if category not in devices_by_category:
                devices_by_category[category] = []
            devices_by_category[category].append(j)

    category_probs = {}
    for category, devices in devices_by_category.items():
        prob_no_fault = np.prod([p_safe[d] for d in devices])
        prob_single_fault = sum(
            p_fail[d] * np.prod([p_safe[other] for other in devices if other != d])
            for d in devices
        )
        category_probs[category] = {
            'no_fault': prob_no_fault,
            'single_fault': prob_single_fault,
            'devices': devices
        }

    p_cond_category = {}
    for j in all_faults:
        if j in op0.G.nodes:
            category = device_category(j)
            cat_info = category_probs[category]
            p_j_fail = p_fail[j]
            other_devices = [d for d in cat_info['devices'] if d != j]
            p_others_safe = np.prod([p_safe[d] for d in other_devices])
            denominator = cat_info['no_fault'] + cat_info['single_fault']
            p_cond_category[j] = (p_j_fail * p_others_safe) / denominator

    category_no_fault_probs = {}
    for category, cat_info in category_probs.items():
        category_no_fault_probs[category] = cat_info['no_fault'] / (cat_info['no_fault'] + cat_info['single_fault'])
    prod_safe_all = np.prod(list(category_no_fault_probs.values()))

    p_cond = {}
    for j in all_faults:
        if j in op0.G.nodes:
            category = device_category(j)
            other_cats_prob = 1.0
            for other_cat in category_no_fault_probs.keys():
                if other_cat != category:
                    other_cats_prob *= category_no_fault_probs[other_cat]
            p_cond[j] = p_cond_category[j] * other_cats_prob

    # [已移除 Excel 写入逻辑]
    sum_risk = 0.0
    sum_loss = 0.0
    sum_over = 0.0
    row_count = 1

    def write_row(data, sheet=None, workbook=None, path=None, accumulators=None):
        """只做累计统计；保留原调用位置但不再写入 Excel。"""
        sum_risk, sum_loss, sum_over, row_count = accumulators
        weight = data.get("w", 0.0)
        sum_risk += data.get("tot", 0.0) * weight
        sum_loss += data.get("lost", 0.0) * weight
        sum_over += data.get("over_penalty", 0.0) * weight
        row_count += 1
        return (sum_risk, sum_loss, sum_over, row_count)

    # ---------- 组合清单 ----------
    single_args = [loc for loc in all_faults if loc in op0.G.nodes]
    double_args = [
        (a, b) for a, b in combinations(all_faults, 2)
        if device_category(a) != device_category(b)
        and a in op0.G.nodes and b in op0.G.nodes
    ]

    # ===== 单一进程池：先单后双（但双故障用“无 L + 外推”的估计法）=====
    with mp.Pool(processes=cpu_num,
                 initializer=init_worker,
                 initargs=(op0.G, p_cond, prod_safe_all)) as pool:

        print(f"启动任务：{len(single_args)} 单故障 + {len(double_args)} 双故障")

        # ---------- 单故障（保持不变） ----------
        cnt = 0
        for result in pool.imap_unordered(simulate_single, single_args, chunksize=1):
            loc, lost, over, tot, strat, overload, w = result
            cnt += 1
            if cnt % PROGRESS_INT == 0:
                print(f"  单故障进度 {cnt}/{len(single_args)}")

            accumulators = write_row({
                "capacity": cap,
                "type": "single",
                "loc": loc,
                "best_strategy": strat,
                "lost": lost,
                "over_penalty": over,
                "tot": tot,
                "overload_lines": ";".join(f"{bid}:{I:.1f}A" for bid, I in overload),
                "w": w
            }, None, None, None, (sum_risk, sum_loss, sum_over, row_count))
            sum_risk, sum_loss, sum_over, row_count = accumulators

            if cnt % PROGRESS_INT == 0:
                print(f"[{cap} kW] E(loss)={sum_loss:,.2f} E(over)={sum_over:,.2f} "
                      f"E(risk)={sum_risk:,.2f}")

        # ---------- 双故障（估计法：仅遍历“不含 L”组合，并按比例外推） ----------
        # 再按固定比例 0.469760170738439 外推到“全部双故障”的贡献。:contentReference[oaicite:2]{index=2}
        ESTIMATE_DOUBLE_RATIO = 0.469760170738439  
        double_args_nol = [(a, b) for (a, b) in double_args
                           if not (a.startswith('L') or b.startswith('L'))]  # :contentReference[oaicite:4]{index=4}

        # 为了打印“估计中的总体期望”，记录单故障结束时的基线
        base_sum_loss, base_sum_over, base_sum_risk = sum_loss, sum_over, sum_risk
        sum_lossw = 0.0   # 仅统计“不含 L”的累加（未外推）
        sum_overw = 0.0
        sum_totw  = 0.0

        print(f"  双故障将仅遍历无 'L' 组合：{len(double_args_nol)} 个，外推比例系数 {ESTIMATE_DOUBLE_RATIO}")

        cnt2 = 0
        for result in pool.imap_unordered(simulate_double, double_args_nol, chunksize=1):
            loc1, loc2, lost, over, tot, strat, w = result
            cnt2 += 1

            # 1) 先累计“未外推”的权重和（用于进度中的近似显示）
            sum_lossw += lost * w
            sum_overw += over * w
            sum_totw  += tot  * w

            # 2) 真正计入总账时：把权重放大为 w / ratio，相当于把“不含 L”的样本外推到“全部双故障”
            w_scaled = w / ESTIMATE_DOUBLE_RATIO
            accumulators = write_row({
                "capacity": cap,
                "type": "double_est",   # 标注：双故障（估计）
                "loc": f"{loc1}+{loc2}",
                "best_strategy": strat,
                "lost": lost,
                "over_penalty": over,
                "tot": tot,
                "overload_lines": "-",
                "w": w_scaled
            }, None, None, None, (sum_risk, sum_loss, sum_over, row_count))
            sum_risk, sum_loss, sum_over, row_count = accumulators

            if cnt2 % PROGRESS_INT == 0 or cnt2 == len(double_args_nol):
                approx_E_over = base_sum_over + (sum_overw / ESTIMATE_DOUBLE_RATIO)
                approx_E_risk = base_sum_risk + (sum_totw  / ESTIMATE_DOUBLE_RATIO)
                print(f"  双故障估计进度 {cnt2}/{len(double_args_nol)} | "
                      f"E(over)≈{approx_E_over:,.2f} E(risk)≈{approx_E_risk:,.2f}") 

    # （收尾汇总打印；不再写Excel）
    print(f"[{cap} kW] E(loss)={sum_loss:,.2f} E(over)={sum_over:,.2f} E(risk)={sum_risk:,.2f}")

# ========= 分段缓存管理器（统一缓存池） =========
class SegmentCacheManager:
    """
    管理分段常值函数的缓存
    缓存仅与DG的实际供电量有关，与时间无关
    """
    def __init__(self, cache_file: Optional[str] = None):
        self.cache: Dict[int, float] = {}  # segment_id -> 函数值（与时间无关）
        self.segments: List[Tuple[np.ndarray, np.ndarray]] = []  # [(lower_bounds, upper_bounds), ...]
        self.cache_file = cache_file
        self.hit_count = 0
        self.miss_count = 0
        
        # 如果指定了缓存文件，尝试加载
        if cache_file and Path(cache_file).exists():
            self.load_cache()
    
    def define_segments_cartesian(self, dim_breakpoints: List[List[float]]):
        """
        通过笛卡尔积方式定义分段区间
        dim_breakpoints: 长度为8的列表，每个元素是该维度的分段点列表
        """
        assert len(dim_breakpoints) == 8, "需要8个维度的分段点"
        
        segments = []
        
        # 为每个维度生成区间 [a, b)
        dim_intervals = []
        for dim_idx, breakpoints in enumerate(dim_breakpoints):
            sorted_points = sorted(breakpoints)
            intervals = []
            
            # 添加第一个区间 (-inf, first_point)
            intervals.append((-np.inf, sorted_points[0]))
            
            # 添加中间区间
            for i in range(len(sorted_points) - 1):
                intervals.append((sorted_points[i], sorted_points[i+1]))
            
            # 添加最后一个区间 [last_point, +inf)
            intervals.append((sorted_points[-1], np.inf))
            
            dim_intervals.append(intervals)
        
        # 生成笛卡尔积
        for interval_combination in itertools.product(*dim_intervals):
            lower = np.array([interval[0] for interval in interval_combination])
            upper = np.array([interval[1] for interval in interval_combination])
            segments.append((lower, upper))
        
        self.segments = segments
        print(f"通过笛卡尔积生成了 {len(segments)} 个分段区间")
        
        # 打印统计信息
        for i, breakpoints in enumerate(dim_breakpoints):
            print(f"  DG{i+1}: {len(breakpoints)} 个分段点 -> {len(breakpoints)+1} 个区间")
    
    def find_segment(self, actual_power: np.ndarray) -> Optional[int]:
        """
        根据实际供电量找到所属的分段区间索引
        actual_power: 8维向量，表示8个DG的实际供电量
        """
        for i, (lower, upper) in enumerate(self.segments):
            # 检查是否在区间 [lower, upper) 内
            if np.all(actual_power >= lower) and np.all(actual_power < upper):
                return i
        return None
    
    def get_cached_value(self, actual_power: np.ndarray) -> Optional[float]:
        """
        尝试从缓存获取值（与时间无关）
        actual_power: 8维向量，DG的实际供电量
        """
        seg_idx = self.find_segment(actual_power)
        if seg_idx is not None:
            if seg_idx in self.cache:
                self.hit_count += 1
                return self.cache[seg_idx]
        return None
    
    def set_cached_value(self, actual_power: np.ndarray, value: float):
        """
        将计算结果存入缓存（与时间无关）
        actual_power: 8维向量，DG的实际供电量
        value: 目标函数值
        """
        seg_idx = self.find_segment(actual_power)
        if seg_idx is not None:
            self.cache[seg_idx] = value
            self.miss_count += 1
    
    def save_cache(self):
        """保存缓存到文件"""
        if self.cache_file:
            with open(self.cache_file, 'wb') as f:
                pickle.dump({
                    'cache': self.cache,
                    'segments': self.segments
                }, f)
            print(f"缓存已保存到 {self.cache_file}，共 {len(self.cache)} 个唯一值")
    
    def load_cache(self):
        """从文件加载缓存"""
        try:
            with open(self.cache_file, 'rb') as f:
                data = pickle.load(f)
                self.cache = data['cache']
                self.segments = data['segments']
                print(f"从 {self.cache_file} 加载了 {len(self.cache)} 个唯一缓存值, {len(self.segments)} 个分段")
        except Exception as e:
            print(f"加载缓存失败: {e}")
    
    def print_stats(self):
        """打印缓存统计信息"""
        total = self.hit_count + self.miss_count
        if total > 0:
            hit_rate = self.hit_count / total * 100
            print(f"\n缓存统计: 命中={self.hit_count}, 未命中={self.miss_count}, 命中率={hit_rate:.1f}%")
            print(f"唯一缓存值数量: {len(self.cache)} / {len(self.segments)} 个可能的分段")
            print(f"缓存覆盖率: {len(self.cache) / len(self.segments) * 100:.1f}%")

# 创建全局缓存管理器（所有时间段共享）
cache_manager = SegmentCacheManager(cache_file="pso_cache.pkl")

# 全局变量，用于在并行进程中访问P_seq
_global_P_seq = None

def set_global_P_seq(P_seq):
    """设置全局P_seq供并行进程使用"""
    global _global_P_seq
    _global_P_seq = P_seq

# ========= 定义各个DG的分段点（用户需要根据实际情况修改） =========
def get_dg_breakpoints():
    """
    定义每个DG（分布式电源）的分段点
    这些分段点应该基于实际供电量的范围来定义
    
    注意：分段点应该覆盖 P_seq + x_t 可能的范围
    """

    dim_breakpoints = [

         [0,120,210,330,540,530,750,1035],
         [0,60,120,180,1035],
         [0,120,200,220,270,350,420,1035],
         [0,40,90,130,190,240,280,1035],
         [0,90,210,280,300,390,400,490,1035],
         [0,350,420,570,620,770,10000],
         [0,200,420,620,680,900,1035],
         [0,150,200,350,400,550,1035],
    ]
    
    
    return dim_breakpoints

# ========= 统一的缓存损失函数 =========
def compute_with_cache(actual_power: np.ndarray) -> float:
    _ensure_segments()  # ← 确保本进/线程已建好分段

    cached_value = cache_manager.get_cached_value(actual_power)
    if cached_value is not None:
        return cached_value

    dg_capacities = list(actual_power)
    risk = process_capacity(dg_capacities)

    # 某些实现会返回 (loss, over, risk)
    if isinstance(risk, (tuple, list)) and len(risk):
        risk = risk[-1]
    if risk is None or (isinstance(risk, float) and not np.isfinite(risk)):
        raise RuntimeError(f"process_capacity 返回无效: {risk}（输入: {dg_capacities}）")

    value = float(risk)
    cache_manager.set_cached_value(actual_power, value)
    return value
def _ensure_segments():
    """如果还没有分段，初始化分段"""
    if not cache_manager.segments:
        # 初始化分段（分段初始化可能很耗时，所以尽量只初始化一次）
        dim_breakpoints = get_dg_breakpoints()
        cache_manager.define_segments_cartesian(dim_breakpoints)


# ========= 损失函数包装器 =========
def loss_fn(t: int, x_t: np.ndarray) -> float:
    """
    损失函数包装器
    t: 时间索引（1-12）
    x_t: 决策变量，形状(8,)
    
    注意：虽然接收t参数（为了保持接口兼容），但实际计算与t无关
    """
    global _global_P_seq
    if _global_P_seq is None:
        # 如果没有设置全局P_seq，直接把x_t作为实际供电量
        return compute_with_cache(x_t)
    
    # 获取时刻t的原始供电量（注意t是从1开始的）
    P_seq_t = _global_P_seq[t - 1]
    
    # 计算实际供电量 = 原始供电量 + 决策变量
    actual_power = P_seq_t + x_t
    
    # 使用统一的缓存计算（与t无关）
    return compute_with_cache(actual_power)

# ========= 单个粒子的目标函数（供并行进程调用，必须顶层定义） =========
def particle_obj(x_single: np.ndarray) -> float:
    # x_single: shape (12, 8)
    # 与原 f(X) 完全一致：逐小时求和（t 从 1 到 12）
    return sum(loss_fn(t + 1, x_single[t]) for t in range(x_single.shape[0]))

# ========= 约束与可行化（保持不变） =========
def build_bounds(P_seq: np.ndarray, Cpv) -> tuple[np.ndarray, np.ndarray]:
    Cpv = np.asarray(Cpv, dtype=float).reshape(1, -1)  # (1,8)
    if Cpv.shape[1] != P_seq.shape[1]:
        raise ValueError("Cpv 维度与 P_seq 列数不一致")
    up_box = 0.15 * Cpv
    lo_box = -0.15 * Cpv
    lo_P   = -np.asarray(P_seq)
    lo = np.maximum(lo_box, lo_P)
    up = np.broadcast_to(up_box, P_seq.shape)
    return lo, up

def project_prefix_feasible(X: np.ndarray) -> np.ndarray:
    T, D = X.shape
    X = X.copy()
    cum = np.zeros(D)
    for t in range(T):
        hi_allowed = -cum
        mask = X[t] > hi_allowed
        X[t, mask] = hi_allowed[mask]
        cum += X[t]
        cum = np.minimum(cum, 0.0)
    return X

def random_feasible(lo: np.ndarray, up: np.ndarray, rng: np.random.Generator) -> np.ndarray:
    T, D = lo.shape
    X = np.zeros((T, D), dtype=float)
    for k in range(D):
        cum = 0.0
        for t in range(T):
            hi_t = min(up[t, k], -cum)
            lo_t = float(lo[t, k])
            if hi_t < lo_t:
                x_tk = hi_t
            else:
                x_tk = rng.uniform(lo_t, hi_t)
            X[t, k] = x_tk
            cum += x_tk
    return X
def print_stats(self):
    """打印缓存统计信息"""
    total = self.hit_count + self.miss_count
    if total > 0:
        hit_rate = self.hit_count / total * 100
        print(f"\n缓存统计: 命中={self.hit_count}, 未命中={self.miss_count}, 命中率={hit_rate:.1f}%")
        print(f"唯一缓存值数量: {len(self.cache)} / {len(self.segments)} 个可能的分段")
        print(f"缓存覆盖率: {len(self.cache) / len(self.segments) * 100:.1f}%")
def _atomic_save_npz(path: Path, **arrays):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")  # 临时文件叫 cap_090.tmp（不带 .npz）

    # 关键：用文件句柄写，Numpy 不会再追加扩展名
    with open(tmp, "wb") as f:
        np.savez_compressed(f, **arrays)

    os.replace(tmp, path)  # 原子替换到 cap_090.npz

# ========= PSO 主体 =========
def pso_optimize(P_seq: np.ndarray,
                 Cpv,
                 swarm_size: int = 50,
                 iters: int = 50,
                 w_start: float = 0.85,
                 w_end: float = 0.35,
                 c1: float = 1.6,
                 c2: float = 1.8,
                 seed: int = 2025,
                 n_jobs: int | None = None,
                 backend: str = "threads",  # 改成 threads
                 ckpt_path: str | None = None,
                 save_every_iter: bool = True,
                 save_on_improve: bool = True
                 ):
    # 其他代码保持不变

    import os, json, time, tempfile
    from pathlib import Path
    import os
    # 设置全局P_seq供损失函数使用
    set_global_P_seq(P_seq)


    def _load_ckpt(path: Path):
        if not path or not Path(path).exists():
            return None
        try:
            data = np.load(path, allow_pickle=True)
            return {k: data[k] for k in data.files}
        except Exception:
            return None

    rng = np.random.default_rng(seed)
    T, D = 12, 8
    if P_seq.shape != (T, D):
        raise ValueError("P_seq 需要形状 (12, 8)")

    lo, up = build_bounds(P_seq, Cpv)

    # 初始化/或从断点恢复
    start_iter = 0
    if ckpt_path:
        ck = _load_ckpt(Path(ckpt_path))
    else:
        ck = None

    if ck is not None:
        try:
            assert tuple(ck["P_seq_shape"]) == tuple(P_seq.shape)
            assert int(ck["swarm_size"]) == int(swarm_size)
            X        = ck["X"];           V        = ck["V"]
            pbest_X  = ck["pbest_X"];     pbest_f  = ck["pbest_f"]
            gbest_X  = ck["gbest_X"];     gbest_f  = float(ck["gbest_f"])
            start_iter = int(ck["iter"]) + 1
            print(f"[{time.strftime('%H:%M:%S')}] 从断点恢复：iter={start_iter}, gbest={gbest_f:.6f}")
        except Exception:
            ck = None

    if ck is None:
        X = np.stack([random_feasible(lo, up, rng) for _ in range(swarm_size)], axis=0)
        V = rng.normal(scale=0.1, size=X.shape)
        pbest_X = X.copy()

        if backend == "threads":
            from concurrent.futures import ThreadPoolExecutor as Executor
        else:
            from concurrent.futures import ProcessPoolExecutor as Executor

        workers = os.cpu_count() - 9
        print(f"[{time.strftime('%H:%M:%S')}] 开始PSO优化，群体大小: {swarm_size}, 迭代次数: {iters}，并行后端: {backend}，并行度: {workers}")

        with Executor(max_workers=workers) as executor:
            pbest_f = np.fromiter(executor.map(particle_obj, pbest_X),
                                  dtype=float, count=swarm_size)
        g_idx   = int(np.argmin(pbest_f))
        gbest_X = pbest_X[g_idx].copy()
        gbest_f = float(pbest_f[g_idx])

        if ckpt_path:
            _atomic_save_npz(Path(ckpt_path),
                             iter=np.array(0, dtype=np.int32),
                             gbest_f=np.array(gbest_f, dtype=float),
                             gbest_X=gbest_X,
                             X=X, V=V,
                             pbest_X=pbest_X, pbest_f=pbest_f,
                             lo=lo, up=up,
                             P_seq=P_seq, Cpv=np.asarray(Cpv, dtype=float),
                             P_seq_shape=np.array(P_seq.shape, dtype=np.int32),
                             swarm_size=np.array(swarm_size, dtype=np.int32),
                             backend=np.array(backend))

    if backend == "threads":
        from concurrent.futures import ThreadPoolExecutor as Executor
    else:
        from concurrent.futures import ProcessPoolExecutor as Executor

    def inertia(it):
        return w_start + (w_end - w_start) * (it / max(1, iters - 1))

    with Executor(max_workers=workers) as executor:
        for it in range(start_iter, iters):
            w = inertia(it)
            r1 = rng.random(size=X.shape)
            r2 = rng.random(size=X.shape)

            X_prev = X
            V = (w * V
                 + c1 * r1 * (pbest_X - X)
                 + c2 * r2 * (gbest_X[np.newaxis, ...] - X))
            X = X + V

            X = np.minimum(np.maximum(X, lo[None, ...]), up[None, ...])
            for i in range(swarm_size):
                X[i] = project_prefix_feasible(X[i])

            vals = np.fromiter(executor.map(particle_obj, X),
                               dtype=float, count=swarm_size)

            improved = vals < pbest_f
            pbest_X[improved] = X[improved]
            pbest_f[improved] = vals[improved]
            gi = int(np.argmin(pbest_f))
            improved_global = pbest_f[gi] < gbest_f
            if improved_global:
                gbest_f = float(pbest_f[gi])
                gbest_X = pbest_X[gi].copy()

            if ckpt_path and (save_every_iter or improved_global):
                _atomic_save_npz(Path(ckpt_path),
                                 iter=np.array(it, dtype=np.int32),
                                 gbest_f=np.array(gbest_f, dtype=float),
                                 gbest_X=gbest_X,
                                 X=X, V=V,
                                 pbest_X=pbest_X, pbest_f=pbest_f,
                                 lo=lo, up=up,
                                 P_seq=P_seq, Cpv=np.asarray(Cpv, dtype=float),
                                 P_seq_shape=np.array(P_seq.shape, dtype=np.int32),
                                 swarm_size=np.array(swarm_size, dtype=np.int32),
                                 backend=np.array(backend))

            if (it + 1) % 10 == 0 or it == 0:
                print(f"[{time.strftime('%H:%M:%S')}] PSO迭代 {it+1:>4}/{iters}, "
                      f"当前最优目标值: {gbest_f:.4f}, 惯性权重: {w:.3f}")
                # 定期打印缓存统计
                if (it + 1) % 50 == 0:
                    cache_manager.print_stats()

    # 打印缓存统计
    cache_manager.print_stats()
    # 保存缓存
    cache_manager.save_cache()

    print(f"[{time.strftime('%H:%M:%S')}] PSO优化完成，最终目标值: {gbest_f:.4f}")
    return gbest_f, gbest_X


if __name__ == "__main__":
    import time
    from pathlib import Path
    import openpyxl

    # ========= 初始化缓存系统（基于笛卡尔积） =========
    print("正在初始化缓存系统...")
    
    # 获取每个DG的分段点（基于实际供电量范围）
    dim_breakpoints = get_dg_breakpoints()
    
    # 使用笛卡尔积方式定义分段
    cache_manager.define_segments_cartesian(dim_breakpoints)
    
    def fmt_sec(s):
        m, s = divmod(s, 60)
        h, m = divmod(int(m), 60)
        return f"{h:d}:{m:02d}:{s:05.2f}"

    def save_result_row(xlsx_path: Path, row: dict, sheet_name: str = "summary"):
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        columns = list(row.keys())
        values  = [row[k] for k in columns]

        if not xlsx_path.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(columns)
            ws.append(values)
            wb.save(xlsx_path)
            wb.close()
        else:
            wb = openpyxl.load_workbook(xlsx_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(columns)
            ws.append(values)
            wb.save(xlsx_path)
            wb.close()

    print(f"\n===== 分布式电源优化程序启动 [{time.strftime('%Y-%m-%d %H:%M:%S')}] =====")

    power_values = [6.21, 68.46, 158.25, 234.78, 282.39, 300.00, 293.04, 267.96, 229.32, 179.16, 119.07, 54.78]
    base_profile = np.array(power_values, dtype=float) / 300.0

    caps = list(range(30, 901, 10))

    pso_kwargs = dict(
        swarm_size=50, iters=10,
        w_start=0.9, w_end=0.35,
        c1=1.7, c2=1.7, seed=2025,
        n_jobs=10,
        backend="threads",
    )

    results_xlsx = Path(__file__).resolve().parent / "结果.xlsx"
    ckpt_dir     = Path(__file__).resolve().parent / "缓存"

    all_results = []

    t0 = time.time()
    for cap in caps:
        ckpt_path = ckpt_dir / f"cap_{cap:03d}.npz"

        print("\n" + "=" * 72)
        print(f"[{time.strftime('%H:%M:%S')}] 开始容量扫描：Cpv = {cap} kW (每个DG相同)")

        Cpv = np.full(8, cap, dtype=float)
        P_seq = np.array([Cpv * base_profile[t] for t in range(len(base_profile))], dtype=float)

        print(f"  - 时间段数量: {len(power_values)}")
        print(f"  - 分布式电源数量: {len(Cpv)}")
        print(f"  - 装机容量向量: {Cpv.tolist()}")

        start_time = time.time()
        best_f, best_X = pso_optimize(
            P_seq, Cpv, **pso_kwargs,
            ckpt_path=str(ckpt_path),     
            save_every_iter=True,         
            save_on_improve=True,
        )
        elapsed = time.time() - start_time

        # 打印实际供电量范围，帮助调整分段点
        actual_power = P_seq + best_X
        print(f"\n实际供电量统计（用于调整分段点）：")
        overall_min = actual_power.min()
        overall_max = actual_power.max()
        print(f"  总体范围: [{overall_min:.1f}, {overall_max:.1f}]")
        for dg_idx in range(8):
            min_val = actual_power[:, dg_idx].min()
            max_val = actual_power[:, dg_idx].max()
            print(f"  DG{dg_idx+1}: [{min_val:.1f}, {max_val:.1f}]")

        X13 = np.zeros(8)
        X_all = np.vstack([best_X, X13])
        print(f"\n[{time.strftime('%H:%M:%S')}] 容量 {cap} kW 结果：")
        print(f"  - 最优目标值: {best_f:.6f}")
        print(f"  - 最优解形状: {best_X.shape}")
        print(f"  - 最优解统计: min={best_X.min():.3f}, max={best_X.max():.3f}, mean={best_X.mean():.3f}")
        print(f"  - 完整解序列形状(含第13行0向量): {X_all.shape}")
        print(f"  - 本轮耗时: {fmt_sec(elapsed)}")

        row = {
            "cap_kW": int(cap),
            "best_f": float(best_f),
            "seconds": float(elapsed),
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "swarm_size": pso_kwargs["swarm_size"],
            "iters": pso_kwargs["iters"],
            "backend": pso_kwargs["backend"],
        }
        save_result_row(results_xlsx, row)
        all_results.append((cap, float(best_f), elapsed))

    total_elapsed = time.time() - t0
    print("\n" + "=" * 72)
    print(f"===== 全部容量扫描完成 [{time.strftime('%Y-%m-%d %H:%M:%S')}]，总耗时: {fmt_sec(total_elapsed)} =====")
    print("扫描汇总（cap_kW, best_f, seconds）：")
    for cap, f, sec in all_results:
        print(f"  - {cap:>3d} kW : best_f = {f:.6f} , time = {sec:.2f}s")
    
    # 最终打印总体缓存统计
    print("\n" + "=" * 72)
    cache_manager.print_stats()